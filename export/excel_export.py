"""
Excel line sheet exporter — clean, standards-compliant .xlsx output.

Hard rules:
- ZERO formulas. Every cell is a static value.
- No merged cells (merge_cells can produce corrupt formula records).
- No print_area formula.
- No pageSetUpPr.
- Simple one-image-per-row anchoring via openpyxl.
- Conservative styling only.
- Must open in Excel for Mac with zero warnings.
"""
from __future__ import annotations

import io
import os
import re
import hashlib
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from utils.schema import EXPORT_COLUMNS, Product
from utils.images import download_image


# ── Styling constants ──
HEADER_FILL = PatternFill(start_color="1A1F2E", end_color="1A1F2E", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1A1F2E")
SUBTITLE_FONT = Font(name="Calibri", size=11, color="666666")
DATE_FONT = Font(name="Calibri", size=10, color="999999")
CELL_FONT = Font(name="Calibri", size=10, color="333333")
LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT_CENTER = Alignment(horizontal="right", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)
ALT_FILL = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")

# Column widths
COL_WIDTHS = {"A": 18, "B": 18, "C": 35, "D": 16, "E": 12, "F": 12, "G": 10, "H": 25}

IMAGE_ROW_HEIGHT = 85  # points


def export_linesheet(
    products: list[Product],
    brand_name: str = "",
    title: str = "Line Sheet",
    date_str: str = "",
    group_by_category: bool = False,
) -> bytes:
    """Export products to a clean .xlsx line sheet. Returns file bytes."""
    wb = openpyxl.Workbook()
    # Remove the default empty sheet's formula cache
    ws = wb.active

    if group_by_category and any(p._category for p in products):
        _build_sheet(ws, products, brand_name, title, date_str, "All Products")
        categories = {}
        for p in products:
            cat = p._category or "Uncategorized"
            categories.setdefault(cat, []).append(p)
        for cat_name, cat_products in categories.items():
            new_ws = wb.create_sheet(title=_safe_sheet_name(cat_name))
            _build_sheet(new_ws, cat_products, brand_name, title, date_str, cat_name)
    else:
        _build_sheet(ws, products, brand_name, title, date_str, "Line Sheet")

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    file_bytes = buf.getvalue()

    # ── Post-export validation ──
    # Reopen and verify it parses without error
    try:
        check_wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        for sheet in check_wb.sheetnames:
            _ = check_wb[sheet].max_row  # Force parse
        check_wb.close()
    except Exception as e:
        raise RuntimeError(f"Export validation failed — generated .xlsx is corrupt: {e}")

    return file_bytes


def _build_sheet(
    ws,
    products: list[Product],
    brand_name: str,
    title: str,
    date_str: str,
    sheet_name: str,
):
    """Build one worksheet with header info + product rows. No formulas, no merges."""
    ws.title = _safe_sheet_name(sheet_name)

    # Column widths
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Row 1: Brand name (plain text in cell A1, no merge) ──
    r = 1
    c = ws.cell(row=r, column=1, value=brand_name or "Brand Name")
    c.font = TITLE_FONT
    c.alignment = LEFT_CENTER
    ws.row_dimensions[r].height = 30

    # ── Row 2: Title ──
    r = 2
    c = ws.cell(row=r, column=1, value=title)
    c.font = SUBTITLE_FONT
    c.alignment = LEFT_CENTER
    ws.row_dimensions[r].height = 20

    # ── Row 3: Date ──
    r = 3
    if not date_str:
        date_str = datetime.now().strftime("%B %d, %Y")
    c = ws.cell(row=r, column=1, value="Date: " + date_str)
    c.font = DATE_FONT
    c.alignment = LEFT_CENTER
    ws.row_dimensions[r].height = 18

    # ── Row 4: spacer ──
    ws.row_dimensions[4].height = 8

    # ── Row 5: Header row ──
    header_row = 5
    for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 28

    # ── Data rows (starting row 6) ──
    data_start = header_row + 1

    for prod_idx, product in enumerate(products):
        data_row = data_start + prod_idx
        ws.row_dimensions[data_row].height = IMAGE_ROW_HEIGHT

        is_alt = prod_idx % 2 == 1

        # Column A: Product Image — try to embed, fallback to URL text
        cell_a = ws.cell(row=data_row, column=1)
        image_embedded = _safe_embed_image(ws, product, data_row)
        if not image_embedded and product.product_image:
            cell_a.value = str(product.product_image)
        cell_a.font = CELL_FONT
        cell_a.alignment = CENTER_CENTER
        cell_a.border = THIN_BORDER
        if is_alt:
            cell_a.fill = ALT_FILL

        # Column B: SKU / UPC — always text format
        cell_b = ws.cell(row=data_row, column=2)
        cell_b.value = str(product.sku_upc) if product.sku_upc else ""
        cell_b.number_format = "@"  # Text format
        cell_b.font = CELL_FONT
        cell_b.alignment = LEFT_CENTER
        cell_b.border = THIN_BORDER
        if is_alt:
            cell_b.fill = ALT_FILL

        # Column C: Product Name
        cell_c = ws.cell(row=data_row, column=3)
        cell_c.value = str(product.product_name) if product.product_name else ""
        cell_c.font = CELL_FONT
        cell_c.alignment = LEFT_CENTER
        cell_c.border = THIN_BORDER
        if is_alt:
            cell_c.fill = ALT_FILL

        # Column D: Color
        cell_d = ws.cell(row=data_row, column=4)
        cell_d.value = str(product.color) if product.color else ""
        cell_d.font = CELL_FONT
        cell_d.alignment = LEFT_CENTER
        cell_d.border = THIN_BORDER
        if is_alt:
            cell_d.fill = ALT_FILL

        # Column E: MSRP
        cell_e = ws.cell(row=data_row, column=5)
        cell_e.value = str(product.msrp) if product.msrp else ""
        cell_e.font = CELL_FONT
        cell_e.alignment = RIGHT_CENTER
        cell_e.border = THIN_BORDER
        if is_alt:
            cell_e.fill = ALT_FILL

        # Column F: Qty
        cell_f = ws.cell(row=data_row, column=6)
        cell_f.value = str(product.qty) if product.qty else ""
        cell_f.font = CELL_FONT
        cell_f.alignment = CENTER_CENTER
        cell_f.border = THIN_BORDER
        if is_alt:
            cell_f.fill = ALT_FILL

        # Column G: Your Cost
        cell_g = ws.cell(row=data_row, column=7)
        cell_g.value = str(product.your_cost) if product.your_cost else ""
        cell_g.font = CELL_FONT
        cell_g.alignment = RIGHT_CENTER
        cell_g.border = THIN_BORDER
        if is_alt:
            cell_g.fill = ALT_FILL

        # Column H: Notes
        cell_h = ws.cell(row=data_row, column=8)
        cell_h.value = str(product.notes) if product.notes else ""
        cell_h.font = CELL_FONT
        cell_h.alignment = LEFT_CENTER
        cell_h.border = THIN_BORDER
        if is_alt:
            cell_h.fill = ALT_FILL

    # Freeze panes below header
    ws.freeze_panes = f"A{header_row + 1}"


def _safe_embed_image(ws, product: Product, row: int) -> bool:
    """
    Embed one product image into cell A of the given row.
    Returns True on success, False on any failure.
    Conservative approach: validate image, convert to clean PNG, simple anchor.
    """
    image_path = product._image_local_path or product.product_image

    if not image_path:
        return False

    image_path = image_path.strip()

    # Download if URL
    if image_path.startswith(("http://", "https://")):
        image_path = download_image(image_path)
    # Amazon-style bare filename (e.g. "41jyPkgADtL.jpg") — try Amazon CDN
    elif re.match(r'^[A-Za-z0-9+\-_]{5,}\.(jpg|jpeg|png|webp)$', image_path):
        image_path = download_image(f"https://m.media-amazon.com/images/I/{image_path}")

    if not image_path or not os.path.exists(image_path):
        return False

    try:
        # Validate the image can be opened
        pil_img = PILImage.open(image_path)
        pil_img.verify()

        # Re-open after verify
        pil_img = PILImage.open(image_path)

        # Convert to RGB if needed (RGBA/P modes can cause issues)
        if pil_img.mode not in ("RGB", "L"):
            pil_img = pil_img.convert("RGB")

        # Resize to fit cell
        pil_img.thumbnail((110, 80), PILImage.LANCZOS)

        # Save to a clean PNG with a unique name per row
        clean_dir = Path(image_path).parent
        clean_name = f"_xl_row{row}.png"
        clean_path = clean_dir / clean_name
        pil_img.save(str(clean_path), "PNG", optimize=False)

        # Create openpyxl image and anchor to cell A{row}
        xl_img = XlImage(str(clean_path))
        xl_img.width = pil_img.width
        xl_img.height = pil_img.height
        ws.add_image(xl_img, f"A{row}")

        return True

    except Exception:
        return False


def _safe_sheet_name(name: str) -> str:
    """Make a string safe for Excel sheet name (max 31 chars, no special chars)."""
    name = re.sub(r'[\\/*?\[\]:]', '', name)
    return name[:31] if name else "Sheet"
